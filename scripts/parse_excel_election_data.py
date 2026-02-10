#!/usr/bin/env python3
"""
Parse Excel files for Ward 27 election results - CORRECTED VERSION
"""

import openpyxl
import json
import os
import re

TENANT_ID = 'bf1a3e36-464e-4eff-b21d-dc71f5a5a582'

def parse_ward_excel(excel_path, ward_suffix):
    """Parse a single Excel file and extract all booth data - FIXED"""
    print(f"\n{'='*60}")
    print(f"Processing: {excel_path}")
    print(f"Ward: 27-{ward_suffix}")
    print(f"{'='*60}")
    
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    
    # Row 9 contains (२७) नवी पठ - पवती designation - booth headers start at column H
    # Find row 9
    header_row = list(sheet.iter_rows(min_row=9, max_row=9, values_only=True))[0]
    
    # Extract booth numbers from odd columns starting from column 7 (H)
    # Pattern: Column 7=booth header, Column 8=booth number
    booth_numbers = []
    col_idx = 7  # Start from column H (index 7)
    
    while col_idx < len(header_row):
        booth_num_cell = header_row[col_idx]
        if booth_num_cell and '.: ' in str(booth_num_cell):
            # Extract number after .: 
            match = re.search(r':\s*(\d+)', str(booth_num_cell))
            if match:
                booth_numbers.append(match.group(1))
        col_idx += 2  # Skip to next booth (every other column)
    
    print(f"Found {len(booth_numbers)} booths: {booth_numbers[:15]}...")
    
    # Extract candidate data from rows 10 onwards
    # Each candidate row has: [idx, ward, candidate_name, party, total, None, vote1, None, vote2, None, ...]
    candidates_data = {}
    candidate_names = []
    
    for row_idx in range(10, min(21, sheet.max_row + 1)):  # Process up to row 20 (candidates + NOTA)
        row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        
        # Get candidate name from columns 2-3 (C-D)
        candidate_parts = []
        for i in [2, 3]:
            if row[i]:
                candidate_parts.append(str(row[i]))
        candidate_name = ' '.join(candidate_parts).strip()
        
        if not candidate_name or 'None' in candidate_name:
            continue
        
        # Check if this is NOTA
        if 'एकही नाही' in candidate_name or 'NOTA' in candidate_name:
            candidate_name = 'NOTA'
        
        # Extract votes from odd columns starting from column 7 (H)
        votes = []
        col_idx = 7
        
        while col_idx < len(row) and len(votes) < len(booth_numbers):
            vote_value = row[col_idx]
            if vote_value is not None and vote_value != '' and vote_value != 'None':
                try:
                    votes.append(int(vote_value))
                except (ValueError, TypeError):
                    votes.append(0)
            else:
                votes.append(0)
            col_idx += 2  # Next booth column
        
        # Ensure we have the right number of votes
        votes = votes[:len(booth_numbers)]
        while len(votes) < len(booth_numbers):
            votes.append(0)
        
        candidates_data[candidate_name] = votes
        candidate_names.append(candidate_name)
        
        total_votes_this_candidate = sum(votes)
        print(f"  {candidate_name[:45]:45} Total: {total_votes_this_candidate:6}, First 5: {votes[:5]}")
    
    wb.close()
    
    # Organize data by booth
    booths = []
    for booth_idx, booth_num in enumerate(booth_numbers):
        booth_votes = {}
        total_votes = 0
        
        for candidate_name in candidate_names:
            if candidate_name in candidates_data and booth_idx < len(candidates_data[candidate_name]):
                vote_count = candidates_data[candidate_name][booth_idx]
                booth_votes[candidate_name] = vote_count
                total_votes += vote_count
        
        if len(booth_votes) > 0 and total_votes > 0:
            # Find winner (excluding NOTA)
            non_nota_votes = {k: v for k, v in booth_votes.items() if k != 'NOTA'}
            if non_nota_votes:
                winner = max(non_nota_votes, key=non_nota_votes.get)
                sorted_votes = sorted(non_nota_votes.values(), reverse=True)
                margin = sorted_votes[0] - sorted_votes[1] if len(sorted_votes) > 1 else 0
            else:
                winner = 'Unknown'
                margin = 0
            
            booths.append({
                'booth_number': booth_num,
                'total_votes': total_votes,
                'candidate_votes': booth_votes,
                'winner': winner,
                'margin': margin
            })
    
    print(f"\n✓ Extracted {len(booths)} booths from Ward 27-{ward_suffix}")
    print(f"  Total votes across all booths: {sum(b['total_votes'] for b in booths)}")
    return booths

def generate_complete_sql(all_ward_data, output_file):
    """Generate complete SQL for all wards"""
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("-- ====================================================================\n")
        f.write("-- Complete Booth-by-Booth Election Results for Ward 27\n")
        f.write(f"-- Tenant: {TENANT_ID}\n")
        f.write("-- Generated from Excel files\n")
        f.write("-- ====================================================================\n\n")
        
        f.write("-- Delete all existing Ward 27 data first\n")
        f.write(f"DELETE FROM election_results WHERE tenant_id = '{TENANT_ID}';\n\n")
        
        total_booths = 0
        total_votes_all = 0
        
        for ward_suffix in ['A', 'B', 'C', 'D']:
            if ward_suffix not in all_ward_data:
                continue
                
            booths = all_ward_data[ward_suffix]
            ward_name = f"Ward 27-{ward_suffix}"
            ward_total_votes = sum(b['total_votes'] for b in booths)
            
            f.write(f"\n-- {'='*60}\n")
            f.write(f"-- Ward 27-{ward_suffix}: {len(booths)} booths, {ward_total_votes:,} total votes\n")
            f.write(f"-- {'='*60}\n\n")
            
            for booth in booths:
                votes_json = json.dumps(booth['candidate_votes'], ensure_ascii=False)
                
                f.write(f"-- Booth {booth['booth_number']}: {booth['total_votes']} votes, Winner: {booth['winner']}\n")
                f.write(f"INSERT INTO election_results (\n")
                f.write(f"    ward_name, booth_number, booth_name,\n")
                f.write(f"    total_voters, total_votes_casted, candidate_votes,\n")
                f.write(f"    winner, margin, tenant_id, created_at\n")
                f.write(f") VALUES (\n")
                f.write(f"    '{ward_name}',\n")
                f.write(f"    '{booth['booth_number']}',\n")
                f.write(f"    'मतदान केंद्र {booth['booth_number']}',\n")
                f.write(f"    0,\n")
                f.write(f"    {booth['total_votes']},\n")
                f.write(f"    '{votes_json}'::jsonb,\n")
                f.write(f"    '{booth['winner']}',\n")
                f.write(f"    {booth['margin']},\n")
                f.write(f"    '{TENANT_ID}',\n")
                f.write(f"    NOW()\n")
                f.write(f");\n\n")
                
                total_booths += 1
                total_votes_all += booth['total_votes']
        
        f.write(f"\n\n-- ====================================================================\n")
        f.write(f"-- Verification Query\n")
        f.write(f"-- ====================================================================\n")
        f.write(f"SELECT \n")
        f.write(f"    ward_name, \n")
        f.write(f"    COUNT(*) as booth_count, \n")
        f.write(f"    SUM(total_votes_casted) as total_votes\n")
        f.write(f"FROM election_results\n")
        f.write(f"WHERE tenant_id = '{TENANT_ID}'\n")
        f.write(f"GROUP BY ward_name\n")
        f.write(f"ORDER BY ward_name;\n")
    
    print(f"\n{'='*70}")
    print(f"✅ SQL GENERATION COMPLETE!")
    print(f"{'='*70}")
    print(f"Output file: {output_file}")
    print(f"Total booths: {total_booths}")
    print(f"Total votes: {total_votes_all:,}")
    print(f"\n📌 Next Steps:")
    print(f"1. Run this SQL file in Supabase SQL Editor")
    print(f"2. Refresh your Election Results page")
    print(f"3. You should see all {total_booths} booths!")

def main():
    base_path = '/Users/shreyasjadhav/Desktop/Nagar 2/Nagarsevak-Managment/result'
    output_sql = '/Users/shreyasjadhav/Desktop/Nagar 2/Nagarsevak-Managment/scripts/import_all_wards_from_excel.sql'
    
    ward_data = {}
    
    for suffix in ['A', 'B', 'C', 'D']:
        excel_file = os.path.join(base_path, f'27{suffix}.xlsx')
        if os.path.exists(excel_file):
            booths = parse_ward_excel(excel_file, suffix)
            if booths:
                ward_data[suffix] = booths
            else:
                print(f"⚠ WARNING: No booths extracted from {excel_file}")
        else:
            print(f"\n⚠ File not found: 27{suffix}.xlsx")
            print(f"  If you have this data, please save it as Excel at: {excel_file}")
    
    if ward_data:
        generate_complete_sql(ward_data, output_sql)
    else:
        print("\n❌ ERROR: No data extracted! Please check the Excel files.")

if __name__ == '__main__':
    main()
